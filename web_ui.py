#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MediaCrawlerPro Web UI
A user-friendly web interface for the MediaCrawler project
"""

import streamlit as st
import asyncio
import json
import sys
import os
from pathlib import Path
import pandas as pd
from datetime import datetime
import httpx
import openpyxl
import subprocess
import threading
import time
import re
import pymysql

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent))

# Set page config
st.set_page_config(
    page_title="MediaCrawlerPro",
    page_icon="🕷️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .status-box {
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    .status-success {
        background-color: #d4edda;
        border-left: 4px solid #28a745;
    }
    .status-warning {
        background-color: #fff3cd;
        border-left: 4px solid #ffc107;
    }
    .post-card {
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        padding: 1rem;
        margin-bottom: 1rem;
        background-color: #ffffff;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .post-card:hover {
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }
    .post-title {
        font-size: 1.2rem;
        font-weight: bold;
        color: #333;
        margin-bottom: 0.5rem;
    }
    .post-meta {
        display: flex;
        gap: 1rem;
        color: #666;
        font-size: 0.9rem;
        margin-top: 0.5rem;
    }
    .post-image {
        max-width: 100%;
        border-radius: 4px;
        margin: 0.5rem 0;
    }
    .status-error {
        background-color: #f8d7da;
        border-left: 4px solid #dc3545;
    }
    .big-number {
        font-size: 2rem;
        font-weight: bold;
        color: #1f77b4;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'crawling' not in st.session_state:
    st.session_state.crawling = False
if 'logs' not in st.session_state:
    st.session_state.logs = []
if 'process' not in st.session_state:
    st.session_state.process = None
if 'last_update' not in st.session_state:
    st.session_state.last_update = time.time()

def check_sign_service():
    """Check if sign service is running"""
    try:
        response = httpx.get("http://localhost:8989/signsrv/pong", timeout=2)
        return response.status_code == 200
    except Exception:
        return False

def check_account_cookies():
    """Check if account cookies are configured"""
    xlsx_path = Path("config/accounts_cookies.xlsx")
    
    if not xlsx_path.exists():
        return False, 0
    
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        
        if "xhs" not in wb.sheetnames:
            return False, 0
        
        sheet = wb["xhs"]
        account_count = 0
        
        for row in range(2, sheet.max_row + 1):
            cookies = sheet.cell(row=row, column=3).value
            if cookies and str(cookies).strip():
                account_count += 1
        
        return account_count > 0, account_count
            
    except Exception:
        return False, 0

def update_config_file(keywords, max_notes, enable_comments, platform, crawler_type):
    """Update the base_config.py file with user settings"""
    config_path = Path("config/base_config.py")
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Update keywords
        content = re.sub(r'KEYWORDS = ".*"', f'KEYWORDS = "{keywords}"', content)
        # Update max notes
        content = re.sub(r'CRAWLER_MAX_NOTES_COUNT = \d+', f'CRAWLER_MAX_NOTES_COUNT = {max_notes}', content)
        # Update comments
        content = re.sub(r'ENABLE_GET_COMMENTS = (True|False)', f'ENABLE_GET_COMMENTS = {enable_comments}', content)
        # Update platform
        content = re.sub(r'PLATFORM = ".*"', f'PLATFORM = "{platform}"', content)
        # Update crawler type
        content = re.sub(r'CRAWLER_TYPE = ".*"', f'CRAWLER_TYPE = "{crawler_type}"', content)
        
        with open(config_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return True
    except Exception as e:
        st.error(f"Error updating config: {e}")
        return False

def load_json_data(platform=None):
    """Load all JSON data files from platform-specific directories"""
    all_data = []
    
    # Platform-specific JSON directories
    platform_paths = {
        "xhs": ["data/xhs/json", "data/xhs"],
        "bili": ["data/bilibili/json", "data/bilibili"],
        "dy": ["data/douyin/json", "data/douyin"],
        "ks": ["data/kuaishou/json", "data/kuaishou"],
        "wb": ["data/weibo/json", "data/weibo"],
        "tieba": ["data/tieba/json", "data/tieba"],
        "zhihu": ["data/zhihu/json", "data/zhihu"],
    }
    
    # If platform specified, only check that platform
    if platform and platform in platform_paths:
        json_paths = platform_paths[platform]
    else:
        # Check all platforms
        json_paths = []
        for paths in platform_paths.values():
            json_paths.extend(paths)
    
    # Add fallback paths
    json_paths.extend([
        Path("data"),  # Data directory (fallback)
        Path("."),  # Root directory (fallback)
    ])
    
    for json_path in json_paths:
        json_path_obj = Path(json_path) if isinstance(json_path, str) else json_path
        if json_path_obj.exists():
            try:
                # Find all JSON files recursively
                data_files = list(json_path_obj.rglob("*.json"))
                for file in data_files:
                    try:
                        with open(file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            if isinstance(data, list):
                                all_data.extend(data)
                            elif isinstance(data, dict):
                                # If it's a single dict, wrap it in a list
                                all_data.append(data)
                    except json.JSONDecodeError as e:
                        # Skip invalid JSON files
                        continue
                    except Exception as e:
                        # Skip files that can't be read
                        continue
            except Exception:
                # Skip paths that can't be accessed
                continue
    
    return all_data

def load_mysql_data():
    """Load data from MySQL database"""
    try:
        connection = pymysql.connect(
            host='localhost',
            port=3307,
            user='root',
            password='123456',
            database='media_crawler',
            charset='utf8mb4'
        )
        
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            # Check if table exists
            cursor.execute("SHOW TABLES LIKE 'xhs_note'")
            if cursor.fetchone():
                cursor.execute("SELECT * FROM xhs_note ORDER BY created_time DESC LIMIT 1000")
                data = cursor.fetchall()
                return data
        connection.close()
        return []
    except Exception as e:
        # Database might not be available or table doesn't exist
        return []

def start_crawler(platform, crawler_type, keywords, max_notes, enable_comments):
    """Start the crawler process"""
    try:
        # Update config file
        if not update_config_file(keywords, max_notes, enable_comments, platform, crawler_type):
            return False
        
        # Try Docker first, fallback to local Python
        try:
            # Stop and remove existing container
            subprocess.run('docker stop mediacrawlerpro', shell=True, capture_output=True, text=True, timeout=5)
            subprocess.run('docker rm mediacrawlerpro', shell=True, capture_output=True, text=True, timeout=5)
            
            # Recreate container with correct platform and type
            project_dir = Path(".").absolute()
            docker_cmd = f'docker run -d --name mediacrawlerpro --network mediacrawlerpro-python_default --restart unless-stopped -v "{project_dir}/config:/app/config" -v "{project_dir}/data:/app/data" -e RELATION_DB_USER=root -e RELATION_DB_HOST=mysql_db -e RELATION_DB_PWD=123456 -e RELATION_DB_PORT=3306 -e RELATION_DB_NAME=media_crawler -e REDIS_DB_HOST=redis_cache -e REDIS_DB_PWD=123456 -e REDIS_DB_PORT=6379 -e REDIS_DB_NUM=0 -e SIGN_SRV_HOST=mediacrawler_signsrv -e SIGN_SRV_PORT=8989 mediacrawlerpro-python-app:latest python main.py --platform {platform} --type {crawler_type}'
            result = subprocess.run(docker_cmd, shell=True, capture_output=True, text=True, timeout=10, cwd=str(project_dir))
            
            if result.returncode != 0:
                # Fallback to docker-compose
                docker_cmd = f'docker-compose run -d --name mediacrawlerpro app python main.py --platform {platform} --type {crawler_type}'
                subprocess.run(docker_cmd, shell=True, capture_output=True, text=True, timeout=10, cwd=str(project_dir))
            
            # Create a dummy process object to track (since Docker runs in background)
            class DockerProcess:
                def __init__(self):
                    self._pid = None
                    self._returncode = None
                
                def poll(self):
                    # Check if container is still running
                    check_cmd = 'docker ps --filter "name=mediacrawlerpro" --format "{{.Status}}"'
                    result = subprocess.run(check_cmd, shell=True, capture_output=True, text=True)
                    if "Up" in result.stdout:
                        return None  # Still running
                    else:
                        # Container exited, check exit code
                        logs_cmd = 'docker ps -a --filter "name=mediacrawlerpro" --format "{{.Status}}"'
                        result = subprocess.run(logs_cmd, shell=True, capture_output=True, text=True)
                        if "Exited (0)" in result.stdout:
                            self._returncode = 0
                        else:
                            self._returncode = 1
                        return self._returncode
                
                @property
                def returncode(self):
                    return self._returncode
            
            return DockerProcess()
        except Exception as e:
            # Fallback to local Python
            env = os.environ.copy()
            env['PYTHONIOENCODING'] = 'utf-8'
            cmd = [sys.executable, 'main.py', '--platform', platform, '--type', crawler_type]
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                env=env,
                cwd=str(Path(".").absolute())
            )
            return process
        
    except Exception as e:
        st.error(f"Error starting crawler: {e}")
        return None

def format_timestamp(timestamp):
    """Format timestamp to readable date"""
    try:
        if isinstance(timestamp, (int, float)) and timestamp > 0:
            dt = datetime.fromtimestamp(timestamp / 1000 if timestamp > 1e10 else timestamp)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(timestamp, str):
            return timestamp
    except:
        pass
    return str(timestamp) if timestamp else "N/A"

def format_user_info(user):
    """Format user information"""
    if isinstance(user, dict):
        return user.get('nickname', user.get('user_name', user.get('user_id', 'Unknown')))
    elif isinstance(user, str):
        return user
    return str(user) if user else "Unknown"

def display_cards(posts):
    """Display posts as cards"""
    for idx, post in enumerate(posts):
        with st.container():
            # Create card container
            col1, col2 = st.columns([3, 1])
            
            with col1:
                # Title
                title = post.get('title') or post.get('desc', '')[:100] or f"Post {idx + 1}"
                if len(title) > 100:
                    title = title[:100] + "..."
                st.markdown(f"### {title}")
                
                # Description
                desc = post.get('desc') or post.get('content', '')
                if desc:
                    if len(desc) > 300:
                        with st.expander("📝 Description"):
                            st.write(desc)
                        st.write(desc[:300] + "...")
                    else:
                        st.write(desc)
                
                # Images
                images = post.get('images', []) or post.get('image_list', [])
                if images:
                    if isinstance(images, list) and len(images) > 0:
                        # Show first image
                        img_url = images[0] if isinstance(images[0], str) else images[0].get('url', '')
                        if img_url:
                            try:
                                st.image(img_url, use_container_width=True, caption="Post Image")
                            except:
                                pass
                        # Show more images if available
                        if len(images) > 1:
                            with st.expander(f"🖼️ View all {len(images)} images"):
                                cols = st.columns(min(3, len(images)))
                                for i, img in enumerate(images[:9]):  # Max 9 images
                                    img_url = img if isinstance(img, str) else img.get('url', '')
                                    if img_url:
                                        try:
                                            cols[i % 3].image(img_url, use_container_width=True)
                                        except:
                                            pass
            
            with col2:
                # Metadata
                st.markdown("**📊 Statistics**")
                
                # Note ID
                note_id = post.get('note_id') or post.get('id', 'N/A')
                st.caption(f"ID: `{str(note_id)[:20]}`")
                
                # Type
                post_type = post.get('type', 'N/A')
                if post_type:
                    st.caption(f"Type: {post_type}")
                
                # User info
                user_info = post.get('user', {})
                user_name = format_user_info(user_info)
                st.caption(f"👤 {user_name}")
                
                # Stats
                col_a, col_b = st.columns(2)
                with col_a:
                    liked = post.get('liked_count', 0) or 0
                    st.metric("❤️", f"{liked:,}" if liked else "0")
                
                with col_b:
                    comments = post.get('comment_count', 0) or 0
                    st.metric("💬", f"{comments:,}" if comments else "0")
                
                # Additional stats
                collected = post.get('collected_count', 0) or 0
                if collected:
                    st.caption(f"⭐ Collected: {collected:,}")
                
                # Timestamp
                created_time = post.get('created_time') or post.get('time', '')
                if created_time:
                    st.caption(f"📅 {format_timestamp(created_time)}")
                
                # Link if available
                note_id_str = str(note_id)
                if note_id_str and note_id_str != 'N/A':
                    st.markdown(f"[🔗 View on XHS](https://www.xiaohongshu.com/explore/{note_id_str})")
            
            # Expandable raw data
            with st.expander("📄 View Raw JSON"):
                st.json(post)
            
            st.markdown("---")

def check_crawler_status():
    """Check Docker container status and return status info"""
    try:
        # Check if container is running
        check_cmd = 'docker ps --filter "name=mediacrawlerpro" --format "{{.Status}}"'
        result = subprocess.run(check_cmd, shell=True, capture_output=True, text=True, timeout=5)
        
        if "Up" in result.stdout:
            return {"status": "running", "message": "🟢 Crawler is running"}
        
        # Check if container exists but stopped
        check_exited_cmd = 'docker ps -a --filter "name=mediacrawlerpro" --format "{{.Status}}"'
        result = subprocess.run(check_exited_cmd, shell=True, capture_output=True, text=True, timeout=5)
        
        if "Exited" in result.stdout:
            # Get last few log lines to see error
            logs_cmd = 'docker-compose logs app --tail=5 2>&1'
            logs_result = subprocess.run(logs_cmd, shell=True, capture_output=True, text=True, timeout=5, cwd=str(Path(".").absolute()))
            logs = logs_result.stdout
            
            # Check for common errors
            error_msg = ""
            if "登录已过期" in logs or "login expired" in logs.lower():
                error_msg = "⚠️ Cookies expired - Please update cookies in accounts_cookies.xlsx"
            elif "账号池中没有可用的账号" in logs:
                error_msg = "⚠️ No available accounts - Please add/update accounts"
            elif "验证码" in logs or "CAPTCHA" in logs:
                error_msg = "⚠️ CAPTCHA detected - Try again later or use IP proxy"
            elif "finished" in logs.lower():
                error_msg = "✅ Crawler finished"
            else:
                error_msg = "⚠️ Crawler stopped"
            
            return {"status": "stopped", "message": error_msg, "logs": logs}
        
        return {"status": "not_found", "message": "⚪ Container not found"}
    except Exception as e:
        return {"status": "error", "message": f"❌ Error checking status: {str(e)}"}

def get_data_stats():
    """Get statistics about collected data"""
    json_data = load_json_data()
    mysql_data = load_mysql_data()
    
    # Count JSON files from all platforms
    json_file_count = 0
    platform_paths = [
        "data/xhs/json", "data/bilibili/json", "data/douyin/json",
        "data/kuaishou/json", "data/weibo/json", "data/tieba/json", "data/zhihu/json",
        "data/xhs", "data/bilibili", "data/douyin", "data/kuaishou",
        "data/weibo", "data/tieba", "data/zhihu", "data", "."
    ]
    for json_path_str in platform_paths:
        json_path = Path(json_path_str)
        if json_path.exists():
            json_file_count += len(list(json_path.rglob("*.json")))
    
    return {
        "total_posts": len(json_data) + len(mysql_data),
        "total_files": json_file_count,
        "json_count": len(json_data),
        "mysql_count": len(mysql_data)
    }

# Header
st.title("🕷️ MediaCrawlerPro Web UI")
st.markdown("---")

# Sidebar - Setup Status
with st.sidebar:
    st.header("📊 System Status")
    
    # Check sign service
    sign_service_ok = check_sign_service()
    if sign_service_ok:
        st.success("✅ Sign Service Running")
    else:
        st.error("❌ Sign Service Not Running")
        st.info("Start the sign service first!")
    
    # Check accounts
    has_accounts, account_count = check_account_cookies()
    if has_accounts:
        st.success(f"✅ {account_count} Account(s) Configured")
    else:
        st.error("❌ No Accounts Configured")
        st.info("Add cookies to accounts_cookies.xlsx")
    
    st.markdown("---")
    
    # Quick stats
    st.header("📈 Quick Stats")
    stats = get_data_stats()
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Posts", stats["total_posts"])
    with col2:
        st.metric("Files", stats["total_files"])
    
    st.markdown("---")
    
    # Setup guide
    if st.button("📖 Setup Guide"):
        with open("SETUP_GUIDE.md", "r", encoding="utf-8") as f:
            st.markdown(f.read())

# Main content tabs
tab1, tab2, tab3, tab4 = st.tabs(["🎮 Control Panel", "📊 Data Viewer", "⚙️ Settings", "📝 Logs"])

with tab1:
    st.header("Control Panel")
    
    # Check if ready to crawl
    can_crawl = sign_service_ok and has_accounts
    
    if not can_crawl:
        st.warning("⚠️ Please complete setup before crawling. Check the sidebar for status.")
        
        with st.expander("🔧 What's needed?"):
            if not sign_service_ok:
                st.error("**Sign Service Not Running**")
                st.code("""# In a new terminal:
cd ..
git clone https://github.com/MediaCrawlerPro/MediaCrawlerPro-SignSrv
cd MediaCrawlerPro-SignSrv
pip install -r requirements.txt
python app.py""", language="bash")
            
            if not has_accounts:
                st.error("**No Account Cookies**")
                st.markdown("""
1. Install [Cookie-Editor](https://chromewebstore.google.com/detail/cookie-editor/hlkenndednhfkekhgcdicdfddnkalmdm)
2. Login to your platform:
   - **XiaoHongShu**: https://www.xiaohongshu.com
   - **Bilibili**: https://www.bilibili.com
   - **Douyin**: https://www.douyin.com
   - **Weibo**: https://m.weibo.cn (H5 page)
   - **Zhihu**: https://www.zhihu.com (use search page cookies)
3. Export cookies (JSON format) and add to `config/accounts_cookies.xlsx` in the corresponding sheet
                """)
    
    st.markdown("---")
    
    # Crawler configuration
    col1, col2 = st.columns(2)
    
    with col1:
        platform = st.selectbox(
            "Platform",
            ["xhs", "dy", "bili", "ks", "wb", "tieba", "zhihu"],
            index=0,
            help="Select the social media platform to crawl"
        )
        
        crawler_type = st.selectbox(
            "Crawler Type",
            ["search", "detail", "creator", "homefeed"],
            index=0,
            help="Type of crawling operation"
        )
    
    with col2:
        keywords = st.text_input(
            "Keywords",
            "deepseek,chatgpt",
            help="Comma-separated keywords for search"
        )
        
        max_notes = st.number_input(
            "Max Posts",
            min_value=1,
            max_value=1000,
            value=40,
            help="Maximum number of posts to crawl"
        )
    
    enable_comments = st.checkbox("Enable Comments", value=True)
    
    st.markdown("---")
    
    # Action buttons
    col1, col2, col3 = st.columns([2, 2, 1])
    
    with col1:
        if st.button("▶️ Start Crawling", disabled=not can_crawl or st.session_state.crawling, use_container_width=True):
            # Update config with user settings
            if update_config_file(keywords, max_notes, enable_comments, platform, crawler_type):
                st.success("✅ Configuration updated!")
                
                # Start crawler
                process = start_crawler(platform, crawler_type, keywords, max_notes, enable_comments)
                if process:
                    st.session_state.process = process
                    st.session_state.crawling = True
                    st.session_state.last_update = time.time()
                    st.success("🚀 Crawler started! Check the 'Data Viewer' tab to see results.")
                    st.info("💡 The crawler is running in the background. Refresh this page to see updates.")
                else:
                    st.error("❌ Failed to start crawler. Check terminal for errors.")
            else:
                st.error("❌ Failed to update configuration file.")
    
    # Show crawler status
    crawler_status = check_crawler_status()
    
    if crawler_status["status"] == "running":
        st.success(crawler_status["message"])
        st.info("💡 The crawler is running in the background. Check the 'Data Viewer' tab for results.")
        # Auto-refresh every 10 seconds when running
        if time.time() - st.session_state.last_update > 10:
            st.session_state.last_update = time.time()
            st.rerun()
    elif crawler_status["status"] == "stopped":
        st.warning(crawler_status["message"])
        if "logs" in crawler_status:
            with st.expander("📋 View Recent Logs"):
                st.text(crawler_status["logs"])
        st.session_state.crawling = False
        st.session_state.process = None
    elif crawler_status["status"] == "not_found":
        st.info(crawler_status["message"])
        st.session_state.crawling = False
        st.session_state.process = None
    
    # Also check session state for legacy support
    if st.session_state.crawling and st.session_state.process:
        if hasattr(st.session_state.process, 'poll'):
            if st.session_state.process.poll() is None:
                st.info("🔄 Crawler is running...")
                if time.time() - st.session_state.last_update > 5:
                    st.session_state.last_update = time.time()
                    st.rerun()
            else:
                st.session_state.crawling = False
                return_code = st.session_state.process.returncode if hasattr(st.session_state.process, 'returncode') else None
                if return_code == 0:
                    st.success("✅ Crawler finished successfully!")
                else:
                    st.error(f"❌ Crawler finished with errors (exit code: {return_code})")
                st.session_state.process = None
    
    with col2:
        if st.button("🔄 Check Status", use_container_width=True):
            st.rerun()
    
    with col3:
        if st.button("🗑️ Clear Data"):
            json_files = list(Path(".").glob("*.json"))
            for file in json_files:
                try:
                    file.unlink()
                except Exception:
                    pass
            st.success("Data cleared!")
            st.rerun()

with tab2:
    st.header("📊 Collected Data")
    
    # Load data from both sources
    json_data = load_json_data()
    mysql_data = load_mysql_data()
    
    # Combine data sources
    all_data = []
    if mysql_data:
        all_data.extend(mysql_data)
    if json_data:
        all_data.extend(json_data)
    
    # Remove duplicates if same note_id exists
    if all_data:
        seen_ids = set()
        unique_data = []
        for item in all_data:
            note_id = item.get('note_id') or item.get('id') or str(item)
            if note_id not in seen_ids:
                seen_ids.add(note_id)
                unique_data.append(item)
        all_data = unique_data
    
    # Show data source info
    col1, col2, col3 = st.columns(3)
    
    # Count JSON files from all platforms
    json_file_count = 0
    platform_paths = [
        "data/xhs/json", "data/bilibili/json", "data/douyin/json",
        "data/kuaishou/json", "data/weibo/json", "data/tieba/json", "data/zhihu/json",
        "data/xhs", "data/bilibili", "data/douyin", "data/kuaishou",
        "data/weibo", "data/tieba", "data/zhihu", "data", "."
    ]
    for json_path_str in platform_paths:
        json_path = Path(json_path_str)
        if json_path.exists():
            json_file_count += len(list(json_path.rglob("*.json")))
    
    with col1:
        st.metric("JSON Files", json_file_count)
        st.caption(f"{len(json_data)} items loaded")
    with col2:
        st.metric("MySQL Records", len(mysql_data))
    with col3:
        st.metric("Total Unique", len(all_data))
    
    if not all_data:
        st.info("📭 No data collected yet. Start crawling to see results here!")
        
        # Show where JSON files should be located
        with st.expander("ℹ️ Where JSON files are saved"):
            platform_info = {
                "xhs": ("data/xhs/json/", "XiaoHongShu"),
                "bili": ("data/bilibili/json/", "Bilibili"),
                "dy": ("data/douyin/json/", "Douyin"),
                "ks": ("data/kuaishou/json/", "Kuaishou"),
                "wb": ("data/weibo/json/", "Weibo"),
                "tieba": ("data/tieba/json/", "Tieba"),
                "zhihu": ("data/zhihu/json/", "Zhihu"),
            }
            
            st.write("JSON files are saved to platform-specific directories:")
            for platform_code, (path, name) in platform_info.items():
                st.write(f"- **{name}** (`{platform_code}`): `{path}`")
            
            st.write("\nFile format: `{crawler_type}_contents_{date}.json`")
            st.write("Example: `search_contents_20251105.json`")
            
            # Check if any directories exist
            found_dirs = []
            for platform_code, (path, name) in platform_info.items():
                json_dir = Path(path)
                if json_dir.exists():
                    files = list(json_dir.glob("*.json"))
                    if files:
                        found_dirs.append((name, path, len(files)))
            
            if found_dirs:
                st.write(f"\n**Found JSON files:**")
                for name, path, count in found_dirs:
                    st.write(f"- {name}: {count} file(s) in `{path}`")
            else:
                st.write("\nNo JSON files found yet. Start crawling to generate data files.")
        
        if st.button("🔄 Refresh Data"):
            st.rerun()
    else:
        st.success(f"✅ Found {len(all_data)} posts!")
        
        # View mode selector
        view_mode = st.radio("View Mode", ["📋 Cards View", "📊 Table View", "📄 JSON View"], horizontal=True)
        
        # Filters and search
        col1, col2, col3 = st.columns([2, 1, 1])
        
        with col1:
            search_term = st.text_input("🔍 Search in posts", "", placeholder="Search by title, description, or content...")
        
        with col2:
            sort_by = st.selectbox("Sort by", ["Newest", "Oldest", "Most Liked", "Most Comments"], index=0)
        
        with col3:
            items_per_page = st.selectbox("Items per page", [10, 20, 50, 100], index=1)
        
        # Filter and sort data
        filtered_data = all_data.copy()
        
        if search_term:
            filtered_data = [
                item for item in filtered_data
                if search_term.lower() in json.dumps(item, ensure_ascii=False).lower()
            ]
        
        # Sort data
        if sort_by == "Newest":
            filtered_data.sort(key=lambda x: x.get('created_time', 0) or 0, reverse=True)
        elif sort_by == "Oldest":
            filtered_data.sort(key=lambda x: x.get('created_time', 0) or 0)
        elif sort_by == "Most Liked":
            filtered_data.sort(key=lambda x: x.get('liked_count', 0) or 0, reverse=True)
        elif sort_by == "Most Comments":
            filtered_data.sort(key=lambda x: x.get('comment_count', 0) or 0, reverse=True)
        
        # Pagination
        total_pages = (len(filtered_data) + items_per_page - 1) // items_per_page
        if total_pages > 1:
            page = st.number_input(f"Page (1-{total_pages})", min_value=1, max_value=total_pages, value=1)
            start_idx = (page - 1) * items_per_page
            end_idx = start_idx + items_per_page
            paginated_data = filtered_data[start_idx:end_idx]
        else:
            paginated_data = filtered_data
            page = 1
        
        st.caption(f"Showing {len(paginated_data)} of {len(filtered_data)} posts (Page {page}/{total_pages if total_pages > 0 else 1})")
        
        if view_mode == "📋 Cards View":
            # Display as cards
            display_cards(paginated_data)
        elif view_mode == "📊 Table View":
            # Display as table
            try:
                df = pd.DataFrame(paginated_data)
                
                # Show important columns
                important_cols = ['note_id', 'title', 'desc', 'type', 'user', 'liked_count', 'collected_count', 'comment_count', 'created_time']
                display_cols = [col for col in important_cols if col in df.columns]
                if display_cols:
                    display_df = df[display_cols].copy()
                else:
                    display_df = df
                
                st.dataframe(display_df, use_container_width=True, height=500)
            except Exception as e:
                st.error(f"Error displaying table: {e}")
                st.json(paginated_data[:5])
        else:
            # JSON View
            st.json(paginated_data)
        
        # Download options
        col1, col2 = st.columns(2)
        with col1:
            try:
                df = pd.DataFrame(filtered_data)
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 Download Filtered Data as CSV",
                    data=csv,
                    file_name=f"mediacrawler_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
            except Exception:
                pass
        
        with col2:
            json_str = json.dumps(filtered_data, ensure_ascii=False, indent=2)
            st.download_button(
                label="📥 Download Filtered Data as JSON",
                data=json_str,
                file_name=f"mediacrawler_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        
        if st.button("🔄 Refresh Data"):
            st.rerun()

with tab3:
    st.header("Settings")
    
    st.subheader("📂 File Paths")
    st.text_input("Project Directory", str(Path(".").absolute()), disabled=True)
    st.text_input("Config Directory", str(Path("config").absolute()), disabled=True)
    
    st.markdown("---")
    
    st.subheader("⚙️ Configuration Files")
    
    config_files = {
        "Base Config": "config/base_config.py",
        "Database Config": "config/db_config.py",
        "Proxy Config": "config/proxy_config.py",
        "Sign Service Config": "config/sign_srv_config.py"
    }
    
    for name, path in config_files.items():
        with st.expander(f"📄 {name}"):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
                st.code(content, language="python")
            except Exception as e:
                st.error(f"Could not read file: {e}")
    
    st.markdown("---")
    
    st.subheader("🔧 Platform-Specific Settings")
    
    import config
    
    info_cols = st.columns(2)
    with info_cols[0]:
        st.info(f"**Current Platform:** {config.PLATFORM}")
        st.info(f"**Crawler Type:** {config.CRAWLER_TYPE}")
        st.info(f"**Keywords:** {config.KEYWORDS}")
    
    with info_cols[1]:
        st.info(f"**Max Posts:** {config.CRAWLER_MAX_NOTES_COUNT}")
        st.info(f"**Storage:** {config.SAVE_DATA_OPTION}")
        st.info(f"**Comments:** {'Enabled' if config.ENABLE_GET_COMMENTS else 'Disabled'}")

with tab4:
    st.header("Logs")
    
    st.info("💡 Logs will appear here when you run the crawler from the terminal.")
    
    # Show recent log files if they exist
    log_files = list(Path("logs").glob("*.log")) if Path("logs").exists() else []
    
    if log_files:
        selected_log = st.selectbox("Select log file", [f.name for f in log_files])
        
        if selected_log:
            log_path = Path("logs") / selected_log
            try:
                with open(log_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    
                # Show last N lines
                num_lines = st.slider("Number of lines to show", 10, 1000, 100)
                
                log_content = "".join(lines[-num_lines:])
                st.text_area("Log Content", log_content, height=400)
                
            except Exception as e:
                st.error(f"Error reading log: {e}")
    else:
        st.info("No log files found yet.")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <p>MediaCrawlerPro Web UI | For educational purposes only | 
    <a href='https://github.com/MediaCrawlerPro/MediaCrawlerPro-Python' target='_blank'>GitHub</a></p>
</div>
""", unsafe_allow_html=True)


