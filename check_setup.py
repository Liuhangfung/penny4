#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Setup verification script for MediaCrawlerPro
Checks if all required components are configured and available
"""

import sys
import httpx
import openpyxl
from pathlib import Path
from colorama import init, Fore, Style

init(autoreset=True)

def print_status(message, success=True):
    """Print a status message with color"""
    icon = "✓" if success else "✗"
    color = Fore.GREEN if success else Fore.RED
    print(f"{color}{icon} {message}{Style.RESET_ALL}")

def check_sign_service():
    """Check if sign service is running"""
    try:
        response = httpx.get("http://localhost:8989/health", timeout=2)
        if response.status_code == 200:
            print_status("Sign service is running on localhost:8989", True)
            return True
    except Exception:
        pass
    
    print_status("Sign service is NOT running on localhost:8989", False)
    print(f"  {Fore.YELLOW}→ You need to start the sign service first")
    print(f"  {Fore.YELLOW}→ See SETUP_GUIDE.md for instructions")
    return False

def check_account_cookies():
    """Check if account cookies are configured"""
    xlsx_path = Path("config/accounts_cookies.xlsx")
    
    if not xlsx_path.exists():
        print_status("accounts_cookies.xlsx not found", False)
        return False
    
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        
        # Check for xhs sheet
        if "xhs" not in wb.sheetnames:
            print_status("No 'xhs' sheet in accounts_cookies.xlsx", False)
            return False
        
        sheet = wb["xhs"]
        
        # Check if there are any accounts (skip header row)
        account_count = 0
        for row in range(2, sheet.max_row + 1):
            cookies = sheet.cell(row=row, column=3).value  # Column C (cookies)
            if cookies and str(cookies).strip():
                account_count += 1
        
        if account_count > 0:
            print_status(f"Found {account_count} account(s) in accounts_cookies.xlsx", True)
            return True
        else:
            print_status("No accounts configured in accounts_cookies.xlsx", False)
            print(f"  {Fore.YELLOW}→ You need to add account cookies")
            print(f"  {Fore.YELLOW}→ See SETUP_GUIDE.md for instructions")
            return False
            
    except Exception as e:
        print_status(f"Error reading accounts_cookies.xlsx: {e}", False)
        return False

def check_config():
    """Check basic configuration"""
    try:
        import config
        
        # Check save data option
        if config.SAVE_DATA_OPTION == "json":
            print_status("Data storage: JSON (no database required)", True)
        elif config.SAVE_DATA_OPTION == "db":
            print_status("Data storage: Database (requires MySQL)", True)
        
        # Check keywords
        if config.KEYWORDS:
            print_status(f"Search keywords: {config.KEYWORDS}", True)
        
        return True
    except Exception as e:
        print_status(f"Error loading config: {e}", False)
        return False

def main():
    print(f"\n{Fore.CYAN}{'='*60}")
    print(f"{Fore.CYAN}MediaCrawlerPro Setup Verification")
    print(f"{Fore.CYAN}{'='*60}\n")
    
    checks = {
        "Configuration": check_config(),
        "Sign Service": check_sign_service(),
        "Account Cookies": check_account_cookies(),
    }
    
    print(f"\n{Fore.CYAN}{'='*60}")
    
    all_passed = all(checks.values())
    
    if all_passed:
        print(f"\n{Fore.GREEN}✓ All checks passed! You're ready to run the crawler.")
        print(f"\n{Fore.CYAN}Run the crawler with:")
        print(f"{Fore.WHITE}  $env:PYTHONIOENCODING=\"utf-8\"; python main.py --platform xhs --type search\n")
    else:
        print(f"\n{Fore.YELLOW}⚠ Some checks failed. Please fix the issues above.")
        print(f"{Fore.CYAN}See SETUP_GUIDE.md for detailed setup instructions.\n")
    
    return 0 if all_passed else 1

if __name__ == "__main__":
    sys.exit(main())


